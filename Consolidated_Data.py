import pandas as pd
import os
import sys
import math
import datetime
import re
import itertools
import openpyxl

class cellPosition:
    row=""
    column=""
    def __init__(self,row,column):
        self.row = row
        self.column = column
              
class DataObj:
    File_Name=""
    Security_Code=""
    Form_Type=""
    Name_of_Acquirer_Seller=""
    Transaction_Date=""
    Reported_to_Exchange_Date=""
    Buy_Sale=""  #Check if buy/sale
    Mode_of_Buy_Sale=""
    Broker_Details=""
    Exchange=""
    Buy_Quantity=""
    Sell_Quantity=""
    Buy_Value=""
    Sell_Value=""
    Buy_Additional_Data=""
    Sell_Additional_Data=""
    
    def __init__(self, File_Name, Security_Code, Form_Type, Name_of_Acquirer_Seller, Transaction_Date, Reported_to_Exchange_Date, Buy_Sale, Mode_of_Buy_Sale, Broker_Details, Exchange, Buy_Quantity, Sell_Quantity, Buy_Value, Sell_Value, Buy_Additional_Data, Sell_Additional_Data):
        self.File_Name=File_Name
        self.Security_Code=Security_Code
        self.Form_Type=Form_Type
        self.Name_of_Acquirer_Seller=Name_of_Acquirer_Seller
        self.Transaction_Date=Transaction_Date
        self.Reported_to_Exchange_Date=Reported_to_Exchange_Date
        self.Buy_Sale=Buy_Sale #Check if buy/sale
        self.Mode_of_Buy_Sale=Mode_of_Buy_Sale
        self.Broker_Details=Broker_Details
        self.Exchange=Exchange
        self.Buy_Quantity=Buy_Quantity
        self.Sell_Quantity=Sell_Quantity
        self.Buy_Value=Buy_Value
        self.Sell_Value=Sell_Value
        self.Buy_Additional_Data=Buy_Additional_Data
        self.Sell_Additional_Data=Sell_Additional_Data
        
class CommentObj:
    Comment_Form_Type=""
    Comment_Name_of_Acquirer_Seller=""
    Comment_Transaction_Date=""
    Comment_Reported_to_Exchange_Date=""
    Comment_Buy_Sale=""
    Comment_Mode_of_Buy_Sale=""
    Comment_Broker_Details=""
    Comment_Exchange=""
    Comment_Buy_Quantity=""
    Comment_Sell_Quantity=""
    Comment_Buy_Value=""
    Comment_Sell_Value=""

    def __init__(self,Comment_Form_Type, Comment_Name_of_Acquirer_Seller, Comment_Transaction_Date, Comment_Reported_to_Exchange_Date, Comment_Buy_Sale, Comment_Mode_of_Buy_Sale, Comment_Broker_Details, Comment_Exchange, Comment_Buy_Quantity, Comment_Sell_Quantity, Comment_Buy_Value, Comment_Sell_Value):
        
        self.Comment_Form_Type=Comment_Form_Type
        self.Comment_Name_of_Acquirer_Seller=Comment_Name_of_Acquirer_Seller
        self.Comment_Transaction_Date=Comment_Transaction_Date
        self.Comment_Reported_to_Exchange_Date=Comment_Reported_to_Exchange_Date
        self.Comment_Buy_Sale=Comment_Buy_Sale #Check if buy/sale
        self.Comment_Mode_of_Buy_Sale=Comment_Mode_of_Buy_Sale
        self.Comment_Broker_Details=Comment_Broker_Details
        self.Comment_Exchange=Comment_Exchange
        self.Comment_Buy_Quantity=Comment_Buy_Quantity
        self.Comment_Sell_Quantity=Comment_Sell_Quantity
        self.Comment_Buy_Value=Comment_Buy_Value
        self.Comment_Sell_Value=Comment_Sell_Value

#############################################################################        
row_num=2
fileLocation="D:\\SampleData\\ConvertedExcel"
#fileLocation="D:\\SamapleData"
bseLocation="D:\\SampleData\\BSE"
nseLocation="D:\\SampleData\\NSE"
path="D:\\DataFile6.xlsx"
system_number=13
sheet_number=1

start_column_for_date=3
end_column_for_date=6
abs_name_from_date=-2
abs_mode_of_sale_from_date=2
abs_trading_member_from_date=4
abs_exchange_from_date=5
abs_buy_quant_from_date=6
abs_buy_value_from_date=7
abs_sell_quant_from_date=8
abs_sell_value_from_date=9


##############################################################################
def hasNumbers(inputString):
    return any(char.isdigit() for char in inputString)


def find_transaction_details(pd_sheet, start_coord, next_date_row, next_date_available):
        row, col =  start_coord
        e_row = 150
        data_string = ''
        comment = ''
        if next_date_available == 1:
            e_row = next_date_row
        for x in range(row, e_row ):
            try:
                if pd_sheet.iloc[x , col]!= None:
                    data_string=data_string+str(pd_sheet.iloc[x , col])
            except:
                comment = 'Data not read properly'
        if len(data_string)>0:
            data_string = re.sub(r'  ','', data_string)
            data_string = re.sub(r'\n', '', data_string)
            data_string = re.sub(r'nan', '', data_string)
        return [data_string , comment]
    


def write_to_delimited_file(data_object, comment_object):
    write_string = ''
    global system_number
    global sheet_number
    for attr, value in data_object.__dict__.items():
        try:
            v = str(value)
            v = v.replace('|', '')
            write_string = write_string + str(v)
            write_string = write_string + str('|')
        except:
            write_string = write_string + str('Error in writing field')
            write_string = write_string + str('|')
    for attr, value in comment_object.__dict__.items():
        try:
            v = str(value)
            v = v.replace('|', '')
            write_string = write_string + str(v)
            write_string = write_string + str('|')
        except:
            write_string = write_string + str('Error in writing field')
            write_string = write_string + str('|')
    write_string = write_string + '\n'
    file = open("C:\\Users\\TradingLab" + str(system_number) + "\\Desktop\\" + str(sheet_number) + '-Converted.txt', 'a')
    file.write(write_string)
    file.close()


def checkDate(file,column):
    date=[]
    for i in range(0,150):
        try:
            try:
                if math.isnan(file.iloc[i,column]):
                    continue
            except:
                if hasNumbers(str(file.iloc[i,column])):
                    if hasNumbers(str(file.iloc[i,column-1])):
                        if not(hasNumbers(str(file.iloc[i,column+1]))):
                            item=cellPosition(i,column-1)
                            date.append(item)
        except:
            break
    return date

def gen_nse_exchange_combos():
    list = [['N', "N'"], ['$', 'S', 'B', 'C', 'Z'],
            ['A', 'E', 'F', 'R', 'T', 'P', 'D', 'G', 'H', 'K', 'L', 'Z', 'C', 'B', 'I:']]
    combinations = ['National', 'tion']
    for element in itertools.product(*list):
        st = ''
        for x in element:
            st = st + x
        combinations.append(st)
    return combinations

def gen_exchange_combos():
    list = [['B', '13', 'I3','E', 'F', 'R', 'P', 'D'], ['$', 'S', 'B', 'Z', 'C'],
            ['A','E', 'F', 'R', 'T', 'P', 'D', 'G', 'H', 'K', 'L', 'Z', 'C', 'B', 'I:']]
    combinations = ['Bombay', 'ombay']
    for element in itertools.product(*list):
        st = ''
        for x in element:
            st = st + x
        combinations.append(st)
    return combinations

def find_exchange(combinations, cell_value_as_str):
    global combination_nse
    cell_value_as_str = re.sub(' ','', cell_value_as_str)
    for x in combinations:
        if x in cell_value_as_str:
            return "BSE"
    for x in combination_nse:
        if x in cell_value_as_str:
            return "NSE"
    return cell_value_as_str

def gen_combos(list):
    combinations = []
    for element in itertools.product(*list):
        st = ''
        for x in element:
            st = st + x
        combinations.append(st)
    return combinations

def form_type(sheet):
    try:
        
        for i in range(len(sheet.index)):
            for j in range(len(sheet.columns)):
                cell_value = str(sheet.iloc[i,j]).replace(" ","")
                cell_value = cell_value.lower()
                #print(cell_value)
                if 'form' in cell_value:
                    cell_value = re.sub('form', '', cell_value)
        
                    if 'c' in cell_value:
                        return 3
                    elif 'd' in cell_value:
                        return 4
                    elif 'a' in cell_value:
                        return 1
                    elif 'b' in cell_value:
                        return 2
                    else:
                        return 100
        return -1
    except IndexError:
        return 10



def get_value(cell_as_str):
    #print("asdasda")
    multiplier = 1
    additional_data = ''
    cell_as_str = cell_as_str.lower()
    list = [['l', '1', 't'],['a','o','c','d'],['c','e', 'k']]
    lakh = gen_combos(list)
    crore = gen_combos([['c', 'e'], [ 'r', 'e', 'i']]) +gen_combos([[ 'c', 'e'], [ 'r', 'e', 'i'] , [ 'o', 'a', 'c', '0'], ['r', 'e', 'i']])
    detect = 0
    #print(cell_as_str)
    for x in lakh:
        if x in cell_as_str:
            detect = 1
            multiplier = 100000
            break
    if detect!=1:
        for x in crore:
            if x in cell_as_str:
                multiplier = 10000000
    cell_as_str = re.sub("\n",'',cell_as_str)
    find_number = re.compile(r'(\d)*((.)(\d))*(\d)*')
    #find_number_1 = re.compile(r'(\d)*(.)*')
    #print(find_number_1.search(cell_as_str))
    number = find_number.search(cell_as_str)
    number = number.group()
    print('asdasdasd')
    cell_as_str = re.sub(r'  ','', cell_as_str)
    cell_as_str = re.sub(number, '', cell_as_str)
    number = float(number)*multiplier
    #print('Im here'+ str(number))
    return [number, str(cell_as_str)]

def getDate(dateVal,File_Name):
    
    monthsAlphaDict={'Jan':'01','January':'01', 'Feb':'02','February':'02', 'Mar':'03', 'March':'03', 'Apr':'04', 'April':'04', 'May':'05', 'Jun':'06', 'June':'06', 'Jul':'07', 'July':'07', 'Aug':'08', 'August':'08', 'Sep':'09', 'September':'09', 'Oct':'10', 'October':'10', 'Nov':'11', 'November':'11', 'Dec':'12', 'December':'12'}
    monthsAlpha=['Jan','January', 'Feb','February', 'Mar', 'March', 'Apr', 'April', 'May', 'Jun', 'June', 'Jul', 'July', 'Aug', 'August', 'Sep', 'September', 'Oct', 'October', 'Nov', 'November', 'Dec', 'December']
    Date=""
    Comment=""
    d_flag=0
    day=""
    month=""
    year=""
    if(isinstance(dateVal, datetime.datetime)):
        if("-" in str(dateVal)):
            d_split=str(dateVal.date()).split("-")
            d_flag=1
    
        elif("/" in str(dateVal)):
            d_split=str(dateVal.date()).split("/")
            d_flag=1
    
        elif("." in str(dateVal)):
            d_split=str(dateVal.date()).split(".")
            d_flag=1
        year,month,day=d_split
        if(year[0:2]=="19"):
            year="20"+year[2:4]
        Date=day+"/"+month+"/"+year
        
        
    elif(isinstance(dateVal, str)):
        count=0
        place=0
        monFlag=0
        if("-" in str(dateVal)):
            d_split=str(dateVal).split("-")
            count=str(dateVal).count("-")
            if(count==2):
                for i in range(len(d_split)):
                    if(d_split[i] in monthsAlpha):
                        place=i
                        monFlag=1
                        break
                if(monFlag==0):
                    day,month,year=d_split
                    d_flag=1
                else:
                    if(place==0):
                        d_flag=1
                        day=d_split[1][0:2]
                        month=monthsAlphaDict[d_split[0]]
                        year=d_split[2]
                    elif(place==1):
                        d_flag=1
                        day=d_split[0][0:2]
                        month=monthsAlphaDict[d_split[1]]
                        year=d_split[2]
                if(len(year)==2):
                    year="20"+year
                if(year[0:2]=="19"):
                    year="20"+year[2:4]
                Date=day+"/"+month+"/"+year
                
        elif("/" in str(dateVal)):
            d_split=str(dateVal).split("/")
            count=str(dateVal).count("/")
            if(count==2):
                for i in range(len(d_split)):
                    if(d_split[i] in monthsAlpha):
                        place=i
                        monFlag=1
                        break
                if(monFlag==0):
                    day,month,year=d_split
                    d_flag=1
                else:
                    if(place==0):
                        d_flag=1
                        day=d_split[1][0:2]
                        month=monthsAlphaDict[d_split[0]]
                        year=d_split[2]
                    elif(place==1):
                        d_flag=1
                        day=d_split[0][0:2]
                        month=monthsAlphaDict[d_split[1]]
                        year=d_split[2]
                if(len(year)==2):
                    year="20"+year
                if(year[0:2]=="19"):
                    year="20"+year[2:4]
                Date=day+"/"+month+"/"+year
    
        elif("." in str(dateVal)):
            d_split=str(dateVal).split(".")
            count=str(dateVal).count(".")
            if(count==2):
                for i in range(len(d_split)):
                    if(d_split[i] in monthsAlpha):
                        place=i
                        monFlag=1
                        break
                if(monFlag==0):
                    day,month,year=d_split
                    d_flag=1
                else:
                    if(place==0):
                        d_flag=1
                        day=d_split[1][0:2]
                        month=monthsAlphaDict[d_split[0]]
                        year=d_split[2]
                    elif(place==1):
                        d_flag=1
                        day=d_split[0][0:2]
                        month=monthsAlphaDict[d_split[1]]
                        year=d_split[2]
                if(len(year)==2):
                    year="20"+year
                if(year[0:2]=="19"):
                    year="20"+year[2:4]
                Date=day+"/"+month+"/"+year
    
        elif(" " in str(dateVal) or "\n" in str(dateVal)):
            d=str(dateVal).replace("\n"," ").replace(","," ").replace("  "," ")
            print("Here: "+ d)
            d_split=d.split(" ")
            if("" in d_split):
                d_split.remove("")
            try:
                for k in range(len(d_split)):
                    if d_split[k] in monthsAlpha:
                        d_flag==2
                        if(k==0 and len(d_split)==3):
                            day=d_split[1][0:2]
                            year=d_split[2]
                            month=monthsAlphaDict[d_split[k]]   
                        elif(k==1 and len(d_split)==3):
                            day=d_split[0][0:2]
                            year=d_split[2]
                            month=monthsAlphaDict[d_split[k]]
                        break
                if(len(year)==2):
                    year="20"+year
                if(year[0:2]=="19"):
                    year="20"+year[2:4]
                Date=day+"/"+month+"/"+year
            except IndexError:
                print("IndexError in "+File_Name)
            
                        
    if(d_flag==0):
        Date=str(dateVal)
        Comment="Format Unexpected"
        
    Date=re.sub('[A-Za-z]', '', Date)
    st=re.search("\d", Date)
    Date=Date[st.start():len(Date)]
    return [Date, Comment]
    
    
def main():
    global fileLocation
    global path
    monthsAlphaDict={'Jan':'01','January':'01', 'Feb':'02','February':'02', 'Mar':'03', 'March':'03', 'Apr':'04', 'April':'04', 'May':'05', 'Jun':'06', 'June':'06', 'Jul':'07', 'July':'07', 'Aug':'08', 'August':'08', 'Sep':'09', 'September':'09', 'Oct':'10', 'October':'10', 'Nov':'11', 'November':'11', 'Dec':'12', 'December':'12'}
    monthsAlpha=['Jan','January', 'Feb','February', 'Mar', 'March', 'Apr', 'April', 'May', 'Jun', 'June', 'Jul', 'July', 'Aug', 'August', 'Sep', 'September', 'Oct', 'October', 'Nov', 'November', 'Dec', 'December']

    columns=["File_Name", "Security_Code", "Form_type", "Name_of_Acquirer_Seller", "Transaction_Date", "Reported_to_Exchange_Date", "Buy_Sale", "Mode_of_Buy_Sale", "Broker_Details","Exchange", "Buy_Quantity", "Sell_Quantity", "Buy_Value", "Sell_Value", "Buy_Additional_Data", "Sell_Additional_Data", "Comment_Form_type", "Comment_Name_of_Acquirer_Seller", "Comment_Transaction_Date", "Comment_Reported_to_Exchange_Date", "Comment_Buy_Sale", "Comment_Mode_of_Buy_Sale", "Comment_Broker_Details", "Comment_Exchange", "Comment_Buy_Quantity", "Comment_Sell_Quantity", "Comment_Buy_Value", "Comment_Sell_Value"]

    
    #################OUTPUT FILE####################
    wb = openpyxl.Workbook()
    wb.save(path)
    wb = openpyxl.load_workbook(path)
    Output_File_Sheet = wb.get_sheet_by_name('Sheet')
    for i in range(1,len(columns)+1):
        Output_File_Sheet.cell(row=1,column=i).value=columns[i-1]
    wb.save(path)
    #############################################
    names=os.listdir(fileLocation)
    
    global start_column_for_date
    global end_column_for_date
    global abs_name_from_date
    global abs_mode_of_sale_from_date
    global abs_trading_member_from_date
    global abs_exchange_from_date
    global abs_buy_quant_from_date
    global abs_buy_value_from_date
    global abs_sell_quant_from_date
    global abs_sell_value_from_date
    
    for i in range(len(names)):
        Security_Code=names[i][0:6]
        File_Name=names[i]
        
        try:
            pandas_file = pd.ExcelFile(fileLocation+"\\"+names[i])
        except:
            print("Can\'t read file")
            data=DataObj(File_Name, Security_Code, "", "", "", "", "", "", "", "", "", "", "", "", "", "")
            comment=CommentObj("", "", "", "", "", "", "", "", "", "", "", "")
            write_to_delimited_file(data, comment)
            continue
        num_of_sheets=pandas_file.sheet_names
        for j in range(len(num_of_sheets)):
            sheet=pandas_file.parse(num_of_sheets[j],header=None)
            F= form_type(sheet)
            Form_Type=""
            
            Comment_Form_Type=""
            if(F==1):
                Form_Type="Form A"
            elif(F==2):
                Form_Type="Form B"
            elif(F==3):
                Form_Type="Form C"
            elif(F==4):
                Form_Type="Form D"
            elif(F==100):
                Comment_Form_Type="Not recognised"
            elif(F==-1):
                Comment_Form_Type="No form type"
            else:
               data=DataObj(File_Name, Security_Code, "", "", "", "", "", "", "", "", "", "", "", "", "", "")
               comment=CommentObj("", "", "", "", "", "", "", "", "", "", "", "")
               addToSheet(wb, Output_File_Sheet, data, comment)
               continue 
            
            location_flag=0
            dates_flag=0
            dates=[]
            for z in range(start_column_for_date,end_column_for_date):
                dates=checkDate(sheet,z)
                if(len(dates)!=0):
                    dates_flag=1
                    break
            print("Dates are "+str(len(dates)))
            f_r=0
            f_c=0
            dates.append("END")
            for x in range(len(dates)-1):
                ###############################Objects to Write to file################################
                Name_of_Acquirer_Seller=""
                Transaction_Date=""
                Reported_to_Exchange_Date=""
                Buy_Sale=""  #Check if buy/sale
                Mode_of_Buy_Sale=""
                Broker_Details=""
                Exchange=""
                Buy_Quantity=""
                Sell_Quantity=""
                Buy_Value=""
                Sell_Value=""
                Buy_Additional_Data=""
                Sell_Additional_Data=""
                
                Comment_Name_of_Acquirer_Seller=""
                Comment_Transaction_Date=""
                Comment_Reported_to_Exchange_Date=""
                Comment_Buy_Sale=""
                Comment_Mode_of_Buy_Sale=""
                Comment_Broker_Details=""
                Comment_Exchange=""
                Comment_Buy_Quantity=""
                Comment_Sell_Quantity=""
                Comment_Buy_Value=""
                Comment_Sell_Value=""
                #######################################################################################
        
                date=dates[x]
                d_flag=0
                if(x==0):
                    f_r=date.row
                    f_c=date.column
                ##################################### Transaction Date ################################
                r=date.row
                c=date.column
                Trans_date=getDate(sheet.iloc[r,c],File_Name)
                Transaction_Date=Trans_date[0]
                Comment_Transaction_Date=Trans_date[1]
                #recheck=getDate(Transaction_Date,File_Name)
                """Transaction_Date=recheck[0]
                Comment_Transaction_Date=recheck[1]"""
                ############################################# Reported Date ###################################
                d_flag=0
                r=date.row
                c=date.column+1
                Report_date=getDate(sheet.iloc[r,c],File_Name)
                Reported_to_Exchange_Date=Report_date[0]
                Comment_Reported_to_Exchange_Date=Report_date[1]
                #recheck=getDate(Reported_to_Exchange_Date,File_Name)
                """Reported_to_Exchange_Date=recheck[0]
                Comment_Reported_to_Exchange_Date=recheck[1]"""
                    ##########################################NAME##############################
                c=c-1
                Name_of_Acquirer_Seller=str(sheet.iloc[r,c+abs_name_from_date]).replace("\n", "")
                if("Name" in Name_of_Acquirer_Seller):
                    Name_of_Acquirer_Seller=str(sheet.iloc[r+1,c+abs_name_from_date]).replace("\n", "")
                if(not(hasNumbers(Name_of_Acquirer_Seller))):
                    Comment_Name_of_Acquirer_Seller="Acceptable"
                else:
                    Comment_Name_of_Acquirer_Seller="Unexpected Characters"


                Mode_of_Buy_Sale=str(sheet.iloc[r,c+abs_mode_of_sale_from_date]).replace("\n", "")
                if(Mode_of_Buy_Sale.isalpha() or (("." in Mode_of_Buy_Sale) and Mode_of_Buy_Sale.isalpha())):
                    Comment_Mode_of_Buy_Sale="Acceptable"
                else:
                    Comment_Mode_of_Buy_Sale="Unexpected Characters"

                try:
                    rp=r
                    trans_det=[]
                    if(dates[x+1]=="END"):
                        trans_det=find_transaction_details(sheet,[rp,c+abs_trading_member_from_date],0,0)
                    else:
                        trans_det=find_transaction_details(sheet,[rp,c+abs_trading_member_from_date],dates[x+1].row,1)
                    
                    Broker_Details = trans_det[0]
                    Comment_Broker_Details=trans_det[1]
                    
                    try:
                        if(not math.isnan(sheet.iloc[r,c+abs_buy_quant_from_date])):
                      
                            buy_quantity=str(sheet.iloc[r,c+abs_buy_quant_from_date])
                            b=get_value(buy_quantity)
                            Buy_Quantity=b[0]
                            Buy_Additional_Data=b[1]
                            print(b[1])
                            Comment_Buy_Quantity="Add present"
                    except:
                       try:
                            buy_quantity=str(sheet.iloc[r,c+abs_buy_quant_from_date])
                            b=get_value(buy_quantity)
                            Buy_Quantity=b[0]
                            Buy_Additional_Data=b[1]
                            print(b[1])
                            Comment_Buy_Quantity="Add present"
                       except:
                            
                             Buy_Quantity="NA"
                             Comment_Buy_Quantity="Unreadable"
                             
                    try:
                        if(not math.isnan(sheet.iloc[r,c+abs_buy_value_from_date])):
                            
                            buy_value=str(sheet.iloc[r,c+abs_buy_value_from_date])
                            b=get_value(buy_value)
                            Buy_Value=b[0]
                            Buy_Additional_Data=b[1]
                            print(b[1])
                            Comment_Buy_Value="Add present"
                    except:
                        try:
                            buy_value=str(sheet.iloc[r,c+abs_buy_value_from_date])
                            b=get_value(buy_value)
                            Buy_Value=b[0]
                            Buy_Additional_Data=b[1]
                            Comment_Buy_Value="Add present"
                        except:
                           
                            Buy_Value="NA"
                            Comment_Buy_Value="Unreadable"

                    try:
                        if(not math.isnan(sheet.iloc[r,c+abs_sell_quant_from_date])):
                            
                            sell_quantity=str(sheet.iloc[r,c+abs_sell_quant_from_date])
                            b=get_value(sell_quantity)
                            Sell_Quantity=b[0]
                            Sell_Additional_Data=b[1]
                            Comment_Sell_Quantity="Add present"
                    except:
                        try:
                            sell_quantity=str(sheet.iloc[r,c+abs_sell_quant_from_date])
                            b=get_value(sell_quantity)
                            Sell_Quantity=b[0]
                            Sell_Additional_Data=b[1]
                            Comment_Sell_Quantity="Add present"
                        except:
                            Sell_Quantity="NA"
                            Comment_Sell_Quantity="Unreadable"
                            
                    try:
                        if(not math.isnan(sheet.iloc[r,c+abs_sell_value_from_date])):
                            
                            sell_value=str(sheet.iloc[r,c+abs_sell_value_from_date])
                            b=get_value(sell_value)
                            Sell_Value=b[0]
                            Sell_Additional_Data=b[1]
                            Comment_Sell_Value="Add present"
                    except:
                        try:
                            sell_value=str(sheet.iloc[r,c+abs_sell_value_from_date])
                            b=get_value(sell_value)
                            Sell_Value=b[0]
                            Sell_Additional_Data=b[1]
                        except:
                            Sell_Value="NA"
                            Comment_Sell_Value="Add present"
                    
                    #Broker_Details=""
                    global combinations
                    Exchange=find_exchange(combinations,str(sheet.iloc[r,c+abs_exchange_from_date]))

                    data=DataObj(File_Name, Security_Code, Form_Type, Name_of_Acquirer_Seller, Transaction_Date, Reported_to_Exchange_Date, Buy_Sale, Mode_of_Buy_Sale, Broker_Details, Exchange, Buy_Quantity, Sell_Quantity, Buy_Value, Sell_Value, Buy_Additional_Data, Sell_Additional_Data)
                    comment=CommentObj(Comment_Form_Type, Comment_Name_of_Acquirer_Seller, Comment_Transaction_Date, Comment_Reported_to_Exchange_Date, Comment_Buy_Sale, Comment_Mode_of_Buy_Sale, Comment_Broker_Details, Comment_Exchange, Comment_Buy_Quantity, Comment_Sell_Quantity, Comment_Buy_Value, Comment_Sell_Value)
                    addToSheet(wb, Output_File_Sheet, data, comment)
                except IndexError:
                    print("IndexError in buy/sell "+File_Name)
                    data=DataObj(File_Name, Security_Code, "", "", "", "", "", "", "", "", "", "", "", "", "", "")
                    comment=CommentObj("", "", "", "", "", "", "", "", "", "", "", "")
                    addToSheet(wb, Output_File_Sheet, data, comment)
                    continue
            
        #for j in range(len(num_of _sheets)):
            #curr=file.parse(num_of_sheets[j])
            #for k in range(0,150):
                #for l in range(0,27):
                 #   if (curr.iloc(k,l) in lis):
                         #location_flag=1

        #if flag==1:
            #os.rename(fileLocation+"\\"+names[i], bseLocation+"\\"+names[i])
        #else:
            #os.rename(fileLocation+"\\"+names[i], nseLocation+"\\"+names[i])
#             
                #print("Hello")"""
combinations=gen_exchange_combos()
combination_nse = gen_nse_exchange_combos()
main()
